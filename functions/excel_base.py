#!/usr/bin/env python3
# -*- coding:utf-8 -*-

import os
import re
import time
import xlrd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def path_name(name, path):
    """
    在当前地址检索目录名或文件名是否有重复，有重复则编号自增
    :param name: 预计文件名
    :return:  最终完整路径
    """
    if name in os.listdir(path):
        fnum = 1
        while True:
            num = '(' + str(fnum) + ')'
            try:
                name_list = list(name)
                idx = name_list.index('.')
                name_list.insert(idx, num)
                new_name = "".join(name_list)
            except:
                new_name = name + num

            if new_name not in os.listdir(path):
                name = new_name
                break
            else:
                fnum += 1
    return path + name


def timer(function):
    """
    装饰器函数timer
    :param function:想要计时的函数
    :return:
    """

    def wrapper(*args, **kwargs):
        time_start = time.time()
        res = function(*args, **kwargs)
        cost_time = time.time() - time_start
        print("【%s】运行时间：【%s】秒" % (function.__name__, cost_time))
        return res

    return wrapper


def get_column_list(column_letter_str, letter_or_num=1):
    """
    对传入的str进行处理：
    1、如果字符串中有分隔符（包括顿号、空格和中英文逗号）则按分隔符将字符串拆成列表，如果没有分隔符，则将按单个元素拆分。
    2、再判断索引是否为字母，若是是字母则按excel列索引规则转换为数字索引。
    :param column_letter_str: 传入的字符串，代表excel列索引
    :param letter_or_num:  传入int或者数字的字符串，则将结果转换为数字索引，否则转换为字母索引
    :return:返回数字索引列表
    """
    if re.search("，|、|,| ", column_letter_str):  # 如果在输入的字符串中找到分隔符，则按分隔符拆分字符串
        column_letter_list = re.split("，|、|,| ", column_letter_str)
    else:  # 如果没有找到分隔符，则按字符拆分字符串
        column_letter_list = [column for column in column_letter_str]

    try:
        if isinstance(letter_or_num, int) or letter_or_num.isdigit():
            # 对输入字符串中的每个元素进行判断处理，如果是字母，则用column_index_from_string()转换为数字索引，否则保留并转为str
            column_index_map = map(lambda x: column_index_from_string(x) if x.isalpha() else int(x), column_letter_list)
        else:
            # 对输入字符串中的每个元素进行判断处理，如果是数字，get_column_letter()转换为字母索引，否则保留并
            column_index_map = map(lambda x: get_column_letter(int(x)) if x.isdigit() else x, column_letter_list)

        column_index_list = [column for column in column_index_map]  # 将上面结果转为list
        return column_index_list
    except ValueError:
        return "如字段中有多位索引，请将所有字段用逗号、顿号或空格隔开。"


# def to_excel(result_data, save_path):
#     """
#     :param result_data: 要输出的数据
#     :param source_file: 源文件名称
#     :param dtype: sheet名称
#     :param save_path: 保存地址
#     :return:
#     """
#     workbook = openpyxl.Workbook()
#     worksheet = workbook.active
#     worksheet.title = dtype
#     for row in result_data:
#         worksheet.append(row)
#
#     for col in range(2, len(result_data[0]) + 1):
#         worksheet.column_dimensions[get_column_letter(col)].width = 20
#
#     source_file_name = os.path.basename(source_file)
#     result_file_name = source_file_name[:source_file_name.index(".")] + "【%s】" % dtype + ".xlsx"
#     result_file_name = path_name(result_file_name, path=save_path)
#     workbook.save(filename=result_file_name)
#     tip_text = "%s输出完成，请查看%s/%s" % (dtype, os.getcwd(), result_file_name)
#     print(tip_text)
#     return tip_text


class ToExcel:
    def __init__(self, source_file, dtype):
        self.source_file = source_file
        self.dtype = dtype

    @property
    def get_default_name(self):
        """根据来源数据文件名和处理方式生成最初的文件名"""
        source_file_name = os.path.basename(self.source_file)
        default_name = source_file_name[:source_file_name.index(".")] + "【%s】" % self.dtype + ".xlsx"
        return default_name

    def get_last_path(self, default_name, save_path):
        """将最初的文件名放到保存路径中,进行重名检测"""
        last_path = path_name(default_name, path=save_path)
        return last_path

    def to_excel(self, result_data, save_path):
        """将数据写入到目标路径，没有重名检测过程，需要手动完成"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = self.dtype
        for row in result_data:
            worksheet.append(row)

        for col in range(2, len(result_data[0]) + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 20

        # 因tkinter会提示重名，在此去掉重名检测
        # last_path = path_name(os.path.basename(save_path), path=os.path.dirname(save_path) + "/")
        try:
            workbook.save(filename=save_path)
            tip_text = "请查看：\n%s" % save_path
            print(tip_text)
            return tip_text
        except FileNotFoundError:
            pass



class Excel:
    """
    将openpyxl和xlrd的读取xlsx、xls的方法整合到一起
    注意：xls只有读取操作，不能写入！

    功能：
    xlsx、xls:最大行、最大列、和读取指定单元格的值
    xlsx:指定单元格写入、保存工作薄
    """

    def __init__(self, path, sheet_number=1):
        """
        :param path: 文件路径
        :param sheet_number: sheet序号，从1开始
        """
        self.__error = "文件格式错误，请重新选择！"
        self.file_path = path
        self.file_type = path.split('.')[1]
        self.file_name = os.path.basename(path)
        if self.file_name.endswith('.xlsx'):
            try:
                self.workbook = openpyxl.load_workbook(path)
                self.worksheet = self.workbook[self.workbook.sheetnames[sheet_number - 1]]
                self.max_row = self.worksheet.max_row
                self.max_column = self.worksheet.max_column
            except IndexError:
                pass
        elif self.file_name.endswith(".xls"):
            try:
                self.workbook = xlrd.open_workbook(path)
                self.worksheet = self.workbook.sheet_by_index(sheet_number - 1)
                self.max_row = self.worksheet.nrows
                self.max_column = self.worksheet.ncols
            except IndexError:
                pass
        else:
            self.workbook = self.__error
            self.worksheet = self.__error

    def cell(self, column, row):
        """获取单元格对象"""
        if self.file_name.endswith('.xlsx'):
            try:
                return self.worksheet.cell(row=row, column=column)
            except AttributeError:
                return None
        elif self.file_name.endswith(".xls"):
            return self.worksheet.cell(row - 1, column - 1)
        else:
            return self.__error

    def value(self, column, row):
        """
        创建实例后直接  实例名.value(列, 行)即可得到指定单元格的值
        :param column: 列
        :param row: 行
        :return: 单元格的值
        """
        if self.file_name.endswith('.xlsx'):
            try:
                return self.worksheet.cell(row=row, column=column).value
            except AttributeError:
                return None
        elif self.file_name.endswith(".xls"):
            return self.worksheet.cell(row - 1, column - 1).value
        else:
            return self.__error

    def _write(self, column, row, value):
        """写入到xlsx文件"""
        if self.file_name.endswith('.xlsx'):
            self.worksheet.cell(row, column, value)
        else:
            return self.__error

    def _save(self, **new_path):
        """
        保存/另存 文件，传入文件名参数则为另存，不传文件名参数则保存源文件
        :param new_path: 文件名参数
        """
        if self.file_name.endswith('.xlsx'):
            if new_path:
                self.workbook.save(filename=new_path)
            else:
                self.workbook.save(self.path)
        else:
            return self.__error

    def row_data_tuple(self, row, columns=""):
        """获取指定一行数据，放入tuple中"""
        if columns == "":
            return tuple([self.value(column, row) for column in range(1, self.max_column + 1)])
        else:
            return tuple([self.value(column, row) for column in get_column_list(columns)])

    def row_data_tuplee(self, row, columns=""):
        if columns == "":
            return tuple(self.worksheet.row_values(row))

    def part_sheet_data_set(self, columns=""):
        """获取指定字段的数据，不能用@property装饰，需要在调用时传参"""
        if columns == "":
            columns = ",".join([str(column) for column in range(1, self.max_column + 1)])
        column_list = get_column_list(columns)

        sheet_data = set()
        for row in range(1, self.max_row + 1):
            r_data = tuple([self.value(column, row) for column in column_list])
            sheet_data.add(r_data)

        return sheet_data

    @property
    def all_sheet_data_set(self):
        """获取整个sheet的全部数据"""
        return self.part_sheet_data_set()


if __name__ == '__main__':
    file_path1 = "//test1.xls"
    file_path2 = "//test3.xlsx"
    excel_i = Excel(file_path2)
    data = excel_i.all_sheet_data_set
    data2 = excel_i.part_sheet_data_set("abcd")
    print(excel_i.row_data_tuple(2, "1234"))
