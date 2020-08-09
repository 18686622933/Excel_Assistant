#!/usr/bin/env python3
# -*- coding:utf-8 -*-


from functions.excel_base import *
import openpyxl


def augment(dir, save_path, title=1):
    """
    将指定目录下的所有excel文件中的数据，去掉表头，顺次保存在新的excel文件中
    :param dir: 指定路径
    :param title: 表头行数
    :param save_path: 保存目录
    :return: 输出成功提示
    """
    if isinstance(dir, list):
        # 判断传入的是否为列表
        excel_list = dir[1:]
        dir = dir[0]
    else:
        # 如果传入的是文件夹在当前目录中的文件列表os.listdir(path)中筛后缀是excel的文件，再转为list，最后在排序
        excel_list = sorted(list(filter(lambda f: f.endswith(".xlsx") or f.endswith(".xls"), os.listdir(dir))))

    # 先把所有数据放到列表中，在一次性写入excel，竟然比遍历同时写入Excel占用内存还少
    augment_result = []
    for t_row in range(1, title + 1):
        augment_result.append((Excel(dir + excel_list[0]).row_data_tuple(t_row)))  # 创建Excel类要用完整的path路径

    for file in excel_list:
        excel = Excel(dir + file)
        for row in range(2, excel.max_row + 1):
            r_data = [excel.value(column, row) for column in range(1, excel.max_column + 1)]
            augment_result.append(r_data)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for r_data in augment_result:
        worksheet.append(r_data)

    # dtype = "合并结果"
    # result_file_name = path_name("%s.xlsx" % dtype, path)
    # workbook.save(filename=result_file_name)
    workbook.save(filename=save_path)
    print("输出完成，请查看%s" % save_path)
    return "请查看：\n%s" % save_path
    # print(u'当前进程的内存使用：%.4f GB' % (psutil.Process(os.getpid()).memory_info().rss / 1024 / 1024 / 1024))


def table_merge(dir, save_path):
    """
    说明：
    1、模板要放在文件名排序的第一位
    2、单元格为空则顺次查找取值
    3、字体红色 = 相当于空值，用于模板表示例
    4、背景黄色 = 累加
    5、背景绿色 = 求平均值
    """

    if isinstance(dir, list):
        # 判断传入的是否为列表
        excel_list = dir[1:]
        dir = dir[0]
    else:
        # 如果传入的是文件夹在当前目录中的文件列表os.listdir(path)中筛后缀是excel的文件，再转为list，最后在排序
        excel_list = sorted(list(filter(lambda f: f.endswith(".xlsx") or f.endswith(".xls"), os.listdir(dir))))

    template_excel = Excel(dir + excel_list[0])
    template_sign = {"empty": {}, "yellow": {}, "green": {}}

    # 将模板表中需要补全的单元格按规则提取放到字典template_sign中， {"empty": {(列索引，行索引):值, }, "yellow": {}, "green": {}}
    for row in range(1, template_excel.max_row + 1):
        for column in range(1, template_excel.max_column + 1):
            cell = template_excel.cell(column, row)
            try:
                color = cell.fill.fgColor.rgb
            except AttributeError:
                color = "00000000"

            try:
                v = cell.value
            except AttributeError:
                v = None

            if color == "FFFFFF00":
                template_sign["yellow"][(column, row)] = 0
            elif color == "FF00B050":
                template_sign["green"][(column, row)] = {"sum": 0, "count": 0}

            elif not (v and color != "FFFF0000"):
                template_sign["empty"][(column, row)] = None

    # 将其他字表不空的数据保存到字典sub_excel_data中
    sub_excel_data = {}
    for file in excel_list[1:]:
        excel = Excel(dir + file)
        one_excel_data = {}
        for row in range(1, excel.max_row + 1):
            for column in range(1, excel.max_column + 1):
                one_excel_data[column, row] = excel.value(column, row)

        sub_excel_data[file] = one_excel_data

    # 统计其他字表数据，更新template_sign
    for f, d in sub_excel_data.items():
        for col_row, value in d.items():
            if value:
                if col_row in template_sign["empty"].keys():
                    template_sign["empty"][col_row] = value
                elif col_row in template_sign["yellow"].keys():
                    template_sign["yellow"][col_row] += value
                elif col_row in template_sign["green"].keys():
                    print(col_row, value)
                    template_sign["green"][col_row]["sum"] += value
                    template_sign["green"][col_row]["count"] += 1

    # 重新打开模板文件，将统计结果写入并保存
    workbook = openpyxl.load_workbook(dir + excel_list[0])
    worksheet = workbook[workbook.sheetnames[0]]
    for col_row, value in template_sign["empty"].items():
        if value:
            worksheet.cell(col_row[1], col_row[0], value)

    for col_row, value in template_sign["yellow"].items():
        if value:
            worksheet.cell(col_row[1], col_row[0], value)
    for col_row, value in template_sign["green"].items():
        if value:
            worksheet.cell(col_row[1], col_row[0], value["sum"] / value["count"])

    # workbook.save(filename=path + excel_list[0])  #保存到模板中
    workbook.save(filename=save_path)  # 另存为
    print("输出完成，请查看%s" % save_path)
    return "请查看：\n%s" % save_path


if __name__ == '__main__':
    file_path1 = "//test_data2/"
    file_path2 = "//test3.xlsx"
    file_path3 = "//test_save_path】.xlsx"
    # augment(file_path1, 1)
    table_merge(file_path1, file_path3)
